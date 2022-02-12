#import nltk
#nltk.download('punkt')
import openpyxl
import re
from nltk.tokenize import word_tokenize
import pandas as pd
from sklearn.linear_model import LogisticRegression
from sklearn.tree import DecisionTreeClassifier
from sklearn.svm import SVC
from sklearn.model_selection import cross_val_score
from sklearn.metrics import f1_score
from sklearn.preprocessing import LabelEncoder
import random
from sklearn.feature_extraction.text import TfidfVectorizer

file1 = openpyxl.load_workbook("tweets.xlsx")
Tweets = file1.active
file2 = openpyxl.load_workbook("filtered tweets.xlsx")
Filtered_tweets = file2.active

def clean(tweet):
    tweet = re.sub("[إأٱآا]", "ا", tweet)
    tweet = re.sub("ٳ", "ء", tweet)
    tweet = re.sub("[ۈؤ]", "و", tweet)
    tweet = re.sub("ئ", "يء", tweet)
    tweet = re.sub("ء", "ئ", tweet)
    tweet = re.sub("ة", "ه", tweet)
    tweet = re.sub("گ", "ك", tweet)
    tweet = re.sub("ﺂ", "", tweet)
    tweet = re.sub("ھ", "ه", tweet)
    tweet = re.sub("[-_#.…;/،)({}*,:!``⇟ღ~↴˺؛'['ϻ?×'↓%ོ؟`٭=%ѕмüı•''—+↩“”|ೋ⏰°‼️۝⇣⌒ ‿ ⌒ٰٰ  ⃣ ⃣ ⏬⏳ ˓͢ ⁦ $ ^»«]", " ", tweet)
    tweet = re.sub("]", " ", tweet)
    tweet = re.sub("[ّ َ ً ُ ٌ ِ ٍ ْـ]", " ", tweet)
    tweet = re.sub(r'(.)\1+', r'\1', tweet)
    tweet = re.sub(r"\d+", "", tweet)
    tweet = re.sub(r"[a-zA-Z]", "", tweet)
    tweet = re.sub("[\U0001F600-\U0001F64F \U0001F300-\U0001F5FF \U0001F680-\U0001F6FF \U0001F1E0-\U0001F1FF \U00002500-\U00002BEF "
                   "\U00002702-\U000027B0 \U00002702-\U000027B0 \U000024C2-\U0001F251 \U0001f926-\U0001f937 \U00010000-\U0010ffff "
                   "\u2640-\u2642 \u2600-\u2B55 \u200d \u23cf \u23e9 \u231a \ufe0f \u3030]", " ", tweet)
    #print(tweet, end=" ") // to see values of filtered_tweets in output screen
    return tweet

def remove_stop_words(tweet):
    list = open('stop_words_arabic.txt', 'r', encoding='utf-8')
    stop_words = list.read().split('\n')
    needed_words = []
    words = word_tokenize(tweet)
    for w in words:
        if w not in (stop_words):
            needed_words.append(w)
    filtered_tweet = " ".join(needed_words)
    list.close()
    return filtered_tweet

def classification(model):
    ft = pd.read_excel("filtered tweets.xlsx")
    encoder = LabelEncoder()
    x = encoder.fit_transform(ft['sentence'])
    x = x.reshape(-1, 1)
    y = encoder.fit_transform(ft['type'])
    random.shuffle(x, random.seed(1.75))
    score = cross_val_score(model, x, y, cv=20, scoring='accuracy').mean()
    print("score : ", score)
    model.fit(x, y)
    y_predict = model.predict(x)
    print(y_predict)
    f_measure = f1_score(y, y_predict)
    print("f-measure : ", f_measure)
    print("--------------------------------------------------------------------------------------")

def features():
    vectorizer = TfidfVectorizer(max_features=100)
    ft = pd.read_excel("filtered tweets.xlsx")
    feature_score = vectorizer.fit_transform(ft['sentence'].values.astype('U'))
    voc = vectorizer.vocabulary_
    print(voc)
    print(feature_score)

#Our main..
print("code is running..:")
for i in range(2, Tweets.max_row + 1):
    #print("\n") // if u used the upper print use this one too, it makes lines appear below each other instead of next to each other
    for j in range(2, Tweets.max_column + 1):
        cell_obj = Tweets.cell(row=i, column=j)
        returned_value = clean(cell_obj.value)
        final_value = remove_stop_words(returned_value)
        Filtered_tweets.cell(row=i, column=j).value = final_value
file2.save('filtered tweets.xlsx')

print("logistic-regission:\n")
logreg = LogisticRegression()
classification(logreg)
print("svm:\n")
svm = SVC(kernel='rbf', random_state=1)
classification(svm)
print("decision-tree:\n")
dectree = DecisionTreeClassifier()
classification(dectree)

features()